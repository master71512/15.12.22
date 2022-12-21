import user_interface as ui
from openpyxl import load_workbook

# def find_name():
#     wb = load_workbook('test.xlsx')
#     name = student_search()
#     students = wb.get_sheet_names()
#     while name not in students:
#         name = wrong_name()
#         if name == '0':
#             goodbye()
#         else:
#             show_scores(name)

def show_scores(name):
    name = ui.find_name()
    w_book = load_workbook('test.xlsx')
    w_sheet = w_book[name]
    print('*\t*\t*\t*\t*\t*\t*')
    for i in range(1, w_sheet.max_row + 1):
        for j in range(1, w_sheet.max_column + 1):
            temp = w_sheet.cell(row=i, column=j)
            print(temp.value, end=' ')
        print()
    print('*\t*\t*\t*\t*\t*\t*')
    ui.goodbye()