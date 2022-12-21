from openpyxl import load_workbook

wb = load_workbook('test.xlsx')
# ws = wb.active
# ws1 = wb.create_sheet('New sheet')
# ws = wb.get_sheet_by_name('Sheet')
ws = wb['Sheet']
print('*\t*\t*\t*\t*\t*\t*')
for i in range(1, ws.max_row + 1):
    for j in range(1, ws.max_column + 1):
        temp = ws.cell(row=i, column=j)
        print(temp.value, end=' ')
    print()
print('*\t*\t*\t*\t*\t*\t*')