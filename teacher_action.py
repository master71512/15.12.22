import user_interface as ui
from openpyxl import load_workbook

def choose_action():
    action = ui.teacher_greeting()
    while action != 0:
        if action == 2:
            add_score()
        if action == 1:
            add_sheet()
        action = ui.teacher_greeting()
    ui.goodbye()

def add_score():
    name = ui.find_name()
    w_book = load_workbook('test.xlsx')
    w_sheet = w_book[name]
    discipline = ui.discipline_enter()
    dis_row = 0
    while dis_row == 0:
        for i in range(1, w_sheet.max_row + 1):
            if w_sheet.cell(row=i, column=1).value == discipline:
                dis_row = i
        if dis_row == 0:
            act = ui.wrong_discipline()
            if act == 1:
                w_sheet.cell(row=w_sheet.max_row + 1, column=1).value = discipline    
                w_book.save('test.xlsx')
                dis_row = w_sheet.max_row
            if act == 2:
                discipline = ui.discipline_enter()
            if act == 0:
                ui.goodbye()
    col_num = 2
    while w_sheet.cell(row=dis_row, column=col_num).value != None:
        col_num += 1
    w_sheet.cell(row=dis_row, column=col_num).value = ui.score_enter()
    w_book.save('test.xlsx')

def add_sheet():
    w_book = load_workbook('test.xlsx')
    name = ui.name_enter()
    w_book.create_sheet(name)
    w_book.save('test.xlsx')